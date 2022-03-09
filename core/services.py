import csv
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ExcelExportService:
    __slots__ = ["queryset", "serializer_class"]

    def __init__(self, queryset, serializer_class):
        self.queryset = queryset
        self.serializer_class = serializer_class

    def prepare_content_data(self):
        data = self.serializer_class(self.queryset, many=True).data
        data = sorted(data, key=lambda k: ((k['get_rating'] and float(k['get_rating']) or 0)), reverse=True)
        header_mappings = self.serializer_class().Meta.field_mappings
        headers = [k for k, v in sorted(header_mappings.items(), key=lambda x: x[1][1])]
        max_column_length = max([len(x[0]) for x in header_mappings.values()])
        return data, headers, header_mappings, max_column_length

    def create_workbook(self):
        wb = Workbook()
        ws = wb.active
        data, headers, header_mappings, length = self.prepare_content_data()
        ws.append([str(header_mappings.get(h)[0]) for h in headers])
        for column in ws.columns:
            ws.column_dimensions[get_column_letter(column[0].column)].width = length
        for row_data in data:
            row = []
            for header in headers:
                row.append(row_data.get(header) or "-")
            ws.append(row)
        return wb

    def create_csv(self, render=True):
        data, headers, header_mappings, length = self.prepare_content_data()

        if render:
            response = HttpResponse('text/csv')
            response['Content-Disposition'] = 'attachment; filename=%s' % self.generate_file_name('csv')
            writer = csv.writer(response)

            writer.writerow([str(header_mappings.get(h)[0]) for h in headers])
            for r in data:
                r['get_rating'] = r['get_rating'] and float(r['get_rating'])
                writer.writerow(list(r.values()))
            return response
        else:
            file_name = self.generate_file_name('csv')
            with open('media/%s' % file_name, 'w') as f:
                writer = csv.writer(f)
                writer.writerow([str(header_mappings.get(h)[0]) for h in headers])
                for r in data:
                    r['get_rating'] = r['get_rating'] and float(r['get_rating'])
                    writer.writerow(list(r.values()))
            f.close()
            return file_name

    def generate_file_name(self, extension):
        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S%f")
        return f"{timestamp}_export.%s" % extension

